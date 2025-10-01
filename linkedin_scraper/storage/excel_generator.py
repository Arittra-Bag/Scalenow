"""Excel file generator for knowledge repository data."""

import pandas as pd
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    Workbook = None
    Font = PatternFill = Alignment = Border = Side = None
    dataframe_to_rows = Table = TableStyleInfo = DataValidation = None

from ..models.knowledge_item import KnowledgeItem, Category
from ..models.exceptions import StorageError
from ..utils.config import Config
from ..utils.logger import get_logger
from .repository_models import KnowledgeRepository

logger = get_logger(__name__)


class ExcelGenerator:
    """Generator for Excel files from knowledge repository data."""
    
    def __init__(self, config: Optional[Config] = None):
        """Initialize the Excel generator."""
        self.config = config or Config.from_env()
        
        if Workbook is None:
            raise StorageError(
                "openpyxl library not installed. "
                "Please install it with: pip install openpyxl"
            )
        
        # Excel styling
        self.styles = {
            'header': {
                'font': Font(bold=True, color='FFFFFF'),
                'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
                'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True)
            },
            'category_ai': PatternFill(start_color='E8F4FD', end_color='E8F4FD', fill_type='solid'),
            'category_saas': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
            'category_marketing': PatternFill(start_color='E8F5E8', end_color='E8F5E8', fill_type='solid'),
            'category_leadership': PatternFill(start_color='F4E8FF', end_color='F4E8FF', fill_type='solid'),
            'category_technology': PatternFill(start_color='FFE8E8', end_color='FFE8E8', fill_type='solid'),
            'category_course': PatternFill(start_color='E8FFE8', end_color='E8FFE8', fill_type='solid'),
            'category_other': PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        logger.info("Excel generator initialized")
    
    def generate_excel_file(
        self,
        repository: KnowledgeRepository,
        output_path: str,
        include_metadata: bool = True,
        include_statistics: bool = True
    ) -> str:
        """Generate a comprehensive Excel file from the knowledge repository."""
        try:
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            # Create workbook
            workbook = Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Create main data sheet
            self._create_main_data_sheet(workbook, repository)
            
            # Create category summary sheet
            if include_statistics:
                self._create_category_summary_sheet(workbook, repository)
            
            # Create metadata sheet
            if include_metadata:
                self._create_metadata_sheet(workbook, repository)
            
            # Create course references sheet
            self._create_course_references_sheet(workbook, repository)
            
            # Save workbook
            workbook.save(output_file)
            
            logger.info(f"Excel file generated: {output_file}")
            return str(output_file)
            
        except Exception as e:
            raise StorageError(f"Failed to generate Excel file: {e}")
    
    def _create_main_data_sheet(self, workbook: Workbook, repository: KnowledgeRepository):
        """Create the main data sheet with all knowledge items."""
        sheet = workbook.create_sheet("Knowledge Repository", 0)
        
        # Define columns
        columns = [
            ('ID', 15),
            ('Topic', 25),
            ('Post Title', 40),
            ('Category', 20),
            ('Key Knowledge Content', 60),
            ('Infographic Summary', 30),
            ('Source Link', 50),
            ('Notes/Applications', 40),
            ('Course References', 30),
            ('Extraction Date', 20)
        ]
        
        # Set column headers and widths
        for col_idx, (header, width) in enumerate(columns, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['border']
            sheet.column_dimensions[cell.column_letter].width = width
        
        # Add data rows
        for row_idx, item in enumerate(repository.items, 2):
            # Prepare data
            course_refs = ', '.join(item.course_references) if item.course_references else ''
            extraction_date = item.extraction_date.strftime('%Y-%m-%d %H:%M') if item.extraction_date else ''
            
            row_data = [
                item.id,
                item.topic,
                item.post_title,
                item.category.value,
                item.key_knowledge_content,
                item.infographic_summary,
                item.source_link,
                item.notes_applications,
                course_refs,
                extraction_date
            ]
            
            # Add data to cells
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.styles['border']
                
                # Apply category-based coloring
                if col_idx == 4:  # Category column
                    cell.fill = self._get_category_fill(item.category)
                
                # Wrap text for content columns
                if col_idx in [3, 5, 7, 8]:  # Title, Content, Notes, Course refs
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                
                # Make source links clickable
                if col_idx == 6 and value:  # Source Link column
                    cell.hyperlink = value
                    cell.font = Font(color='0000FF', underline='single')
        
        # Create table
        if repository.items:
            table_range = f"A1:{sheet.cell(row=len(repository.items) + 1, column=len(columns)).coordinate}"
            table = Table(displayName="KnowledgeTable", ref=table_range)
            
            # Add table style
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            table.tableStyleInfo = style
            sheet.add_table(table)
        
        # Freeze header row
        sheet.freeze_panes = "A2"
        
        # Add data validation for category column
        category_validation = DataValidation(
            type="list",
            formula1=f'"{",".join([cat.value for cat in Category])}"',
            allow_blank=False
        )
        category_validation.error = "Please select a valid category"
        category_validation.errorTitle = "Invalid Category"
        
        if repository.items:
            category_range = f"D2:D{len(repository.items) + 1}"
            sheet.add_data_validation(category_validation)
            category_validation.add(category_range)
    
    def _create_category_summary_sheet(self, workbook: Workbook, repository: KnowledgeRepository):
        """Create a summary sheet with category statistics."""
        sheet = workbook.create_sheet("Category Summary")
        
        # Calculate category statistics
        category_stats = {}
        for item in repository.items:
            category = item.category.value
            if category not in category_stats:
                category_stats[category] = {
                    'count': 0,
                    'recent_items': [],
                    'top_topics': []
                }
            
            category_stats[category]['count'] += 1
            
            # Track recent items (last 10)
            if len(category_stats[category]['recent_items']) < 10:
                category_stats[category]['recent_items'].append({
                    'title': item.post_title,
                    'date': item.extraction_date.strftime('%Y-%m-%d') if item.extraction_date else '',
                    'topic': item.topic
                })
        
        # Headers
        headers = ['Category', 'Count', 'Percentage', 'Recent Items', 'Sample Topics']
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['border']
        
        # Set column widths
        sheet.column_dimensions['A'].width = 25
        sheet.column_dimensions['B'].width = 10
        sheet.column_dimensions['C'].width = 12
        sheet.column_dimensions['D'].width = 50
        sheet.column_dimensions['E'].width = 40
        
        # Add data
        total_items = len(repository.items)
        row_idx = 2
        
        for category, stats in sorted(category_stats.items(), key=lambda x: x[1]['count'], reverse=True):
            percentage = (stats['count'] / total_items * 100) if total_items > 0 else 0
            
            # Recent items summary
            recent_summary = '; '.join([
                f"{item['title'][:30]}..." if len(item['title']) > 30 else item['title']
                for item in stats['recent_items'][:3]
            ])
            
            # Sample topics
            topics = list(set([item['topic'] for item in stats['recent_items']]))[:3]
            topics_summary = '; '.join(topics)
            
            row_data = [
                category,
                stats['count'],
                f"{percentage:.1f}%",
                recent_summary,
                topics_summary
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.styles['border']
                
                # Apply category coloring
                if col_idx == 1:
                    try:
                        cat_enum = Category.from_string(category)
                        cell.fill = self._get_category_fill(cat_enum)
                    except:
                        pass
                
                # Wrap text for longer columns
                if col_idx in [4, 5]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            row_idx += 1
        
        # Add total row
        total_row = row_idx + 1
        sheet.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        sheet.cell(row=total_row, column=2, value=total_items).font = Font(bold=True)
        sheet.cell(row=total_row, column=3, value="100.0%").font = Font(bold=True)
        
        # Freeze header
        sheet.freeze_panes = "A2"
    
    def _create_metadata_sheet(self, workbook: Workbook, repository: KnowledgeRepository):
        """Create a metadata sheet with repository information."""
        sheet = workbook.create_sheet("Metadata")
        
        # Repository metadata
        metadata_items = [
            ('Repository Version', repository.version),
            ('Created At', repository.created_at.strftime('%Y-%m-%d %H:%M:%S')),
            ('Last Updated', repository.updated_at.strftime('%Y-%m-%d %H:%M:%S')),
            ('Total Items', len(repository.items)),
            ('', ''),  # Spacer
            ('Generation Info', ''),
            ('Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            ('Generator Version', '1.0'),
            ('', ''),  # Spacer
        ]
        
        # Add repository metadata
        if repository.metadata:
            metadata_items.append(('Repository Statistics', ''))
            for key, value in repository.metadata.items():
                if isinstance(value, dict):
                    metadata_items.append((f"{key.title()}", ''))
                    for sub_key, sub_value in value.items():
                        metadata_items.append((f"  {sub_key}", str(sub_value)))
                else:
                    metadata_items.append((key.title(), str(value)))
        
        # Headers
        sheet.cell(row=1, column=1, value="Property").font = Font(bold=True)
        sheet.cell(row=1, column=2, value="Value").font = Font(bold=True)
        
        # Set column widths
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 50
        
        # Add metadata
        for row_idx, (prop, value) in enumerate(metadata_items, 2):
            sheet.cell(row=row_idx, column=1, value=prop)
            sheet.cell(row=row_idx, column=2, value=value)
            
            # Style section headers
            if value == '' and prop != '':
                sheet.cell(row=row_idx, column=1).font = Font(bold=True, color='366092')
    
    def _create_course_references_sheet(self, workbook: Workbook, repository: KnowledgeRepository):
        """Create a sheet with all course references."""
        sheet = workbook.create_sheet("Course References")
        
        # Collect all course references
        course_data = []
        for item in repository.items:
            if item.course_references:
                for course in item.course_references:
                    course_data.append({
                        'course': course,
                        'category': item.category.value,
                        'topic': item.topic,
                        'post_title': item.post_title,
                        'source_link': item.source_link,
                        'extraction_date': item.extraction_date.strftime('%Y-%m-%d') if item.extraction_date else ''
                    })
        
        if not course_data:
            sheet.cell(row=1, column=1, value="No course references found in the repository.")
            return
        
        # Headers
        headers = ['Course/Training', 'Category', 'Topic', 'Source Post', 'Source Link', 'Date']
        for col_idx, header in enumerate(headers, 1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.font = self.styles['header']['font']
            cell.fill = self.styles['header']['fill']
            cell.alignment = self.styles['header']['alignment']
            cell.border = self.styles['border']
        
        # Set column widths
        widths = [40, 20, 25, 40, 50, 15]
        for col_idx, width in enumerate(widths, 1):
            sheet.column_dimensions[chr(64 + col_idx)].width = width
        
        # Add data
        for row_idx, course_info in enumerate(course_data, 2):
            row_data = [
                course_info['course'],
                course_info['category'],
                course_info['topic'],
                course_info['post_title'],
                course_info['source_link'],
                course_info['extraction_date']
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.styles['border']
                
                # Make source links clickable
                if col_idx == 5 and value:
                    cell.hyperlink = value
                    cell.font = Font(color='0000FF', underline='single')
                
                # Wrap text for longer columns
                if col_idx in [1, 4]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Freeze header
        sheet.freeze_panes = "A2"
    
    def _get_category_fill(self, category: Category) -> PatternFill:
        """Get fill color for a category."""
        category_fills = {
            Category.AI_MACHINE_LEARNING: self.styles['category_ai'],
            Category.SAAS_BUSINESS: self.styles['category_saas'],
            Category.MARKETING_SALES: self.styles['category_marketing'],
            Category.LEADERSHIP_MANAGEMENT: self.styles['category_leadership'],
            Category.TECHNOLOGY_TRENDS: self.styles['category_technology'],
            Category.COURSE_CONTENT: self.styles['category_course'],
            Category.OTHER: self.styles['category_other']
        }
        
        return category_fills.get(category, self.styles['category_other'])
    
    def generate_simple_excel(
        self,
        repository: KnowledgeRepository,
        output_path: str
    ) -> str:
        """Generate a simple Excel file using pandas (fallback method)."""
        try:
            # Convert to DataFrame
            data = []
            for item in repository.items:
                data.append({
                    'ID': item.id,
                    'Topic': item.topic,
                    'Post Title': item.post_title,
                    'Category': item.category.value,
                    'Key Knowledge Content': item.key_knowledge_content,
                    'Infographic Summary': item.infographic_summary,
                    'Source Link': item.source_link,
                    'Notes/Applications': item.notes_applications,
                    'Course References': ', '.join(item.course_references) if item.course_references else '',
                    'Extraction Date': item.extraction_date.strftime('%Y-%m-%d %H:%M') if item.extraction_date else ''
                })
            
            df = pd.DataFrame(data)
            
            # Save to Excel
            output_file = Path(output_path)
            output_file.parent.mkdir(parents=True, exist_ok=True)
            
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Knowledge Repository', index=False)
                
                # Add category summary
                category_summary = df['Category'].value_counts().reset_index()
                category_summary.columns = ['Category', 'Count']
                category_summary['Percentage'] = (category_summary['Count'] / len(df) * 100).round(1)
                category_summary.to_excel(writer, sheet_name='Category Summary', index=False)
            
            logger.info(f"Simple Excel file generated: {output_file}")
            return str(output_file)
            
        except Exception as e:
            raise StorageError(f"Failed to generate simple Excel file: {e}")
    
    def update_existing_excel(
        self,
        excel_path: str,
        new_items: List[KnowledgeItem]
    ) -> str:
        """Update an existing Excel file with new knowledge items."""
        try:
            excel_file = Path(excel_path)
            if not excel_file.exists():
                raise StorageError(f"Excel file not found: {excel_path}")
            
            # Load existing data
            existing_df = pd.read_excel(excel_path, sheet_name='Knowledge Repository')
            
            # Convert new items to DataFrame
            new_data = []
            for item in new_items:
                new_data.append({
                    'ID': item.id,
                    'Topic': item.topic,
                    'Post Title': item.post_title,
                    'Category': item.category.value,
                    'Key Knowledge Content': item.key_knowledge_content,
                    'Infographic Summary': item.infographic_summary,
                    'Source Link': item.source_link,
                    'Notes/Applications': item.notes_applications,
                    'Course References': ', '.join(item.course_references) if item.course_references else '',
                    'Extraction Date': item.extraction_date.strftime('%Y-%m-%d %H:%M') if item.extraction_date else ''
                })
            
            new_df = pd.DataFrame(new_data)
            
            # Combine data (remove duplicates based on Source Link)
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
            combined_df = combined_df.drop_duplicates(subset=['Source Link'], keep='last')
            
            # Save updated file
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                combined_df.to_excel(writer, sheet_name='Knowledge Repository', index=False)
            
            logger.info(f"Excel file updated with {len(new_items)} new items: {excel_path}")
            return str(excel_path)
            
        except Exception as e:
            raise StorageError(f"Failed to update Excel file: {e}")
    
    def get_next_version_filename(self, base_path: str) -> str:
        """Get the next version filename for Excel files."""
        base_path = Path(base_path)
        base_name = base_path.stem
        extension = base_path.suffix
        directory = base_path.parent
        
        # Find existing versions
        version_pattern = re.compile(rf"{re.escape(base_name)}_v(\d+){re.escape(extension)}")
        versions = []
        
        for file in directory.glob(f"{base_name}_v*{extension}"):
            match = version_pattern.match(file.name)
            if match:
                versions.append(int(match.group(1)))
        
        # Get next version number
        next_version = max(versions, default=0) + 1
        
        return str(directory / f"{base_name}_v{next_version}{extension}")